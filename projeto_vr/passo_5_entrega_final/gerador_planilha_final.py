#!/usr/bin/env python3
"""
Gerador de Planilha Final - Passo 5
Gera a planilha final no formato exato da operadora para envio,
seguindo o modelo "VR MENSAL 05.2025.xlsx".
"""

import csv
import json
import os
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

# Importar sistema de logging
sys.path.append(str(Path(__file__).parent.parent.parent))
from utils.logging_config import log_fim_passo, log_inicio_passo, setup_logging

logger = setup_logging()


class GeradorPlanilhaFinal:
    """Gera planilha final para envio à operadora."""

    def __init__(self):
        self.diretorio_output = self._encontrar_diretorio_output()
        self.custo_empresa_percentual = 0.80
        self.custo_funcionario_percentual = 0.20
        self.tabela_dias_uteis = self._carregar_tabela_dias_uteis()
        self.dias_uteis_por_sindicato = self._carregar_dias_uteis()
        self.valores_vr_sindicato = self._carregar_valores_vr_sindicato()
        
    def _encontrar_diretorio_output(self) -> Path:
        """Garante que o diretório output seja criado na raiz do projeto, independente do nome do diretório."""
        # Assume que este arquivo está em projeto_vr/passo_5_entrega_final/
        raiz_projeto = Path(__file__).resolve().parents[2]
        output_dir = raiz_projeto / "output"
        output_dir.mkdir(exist_ok=True)
        return output_dir
        
    def _carregar_valores_vr_sindicato(self) -> Dict[str, Decimal]:
        """Carrega os valores de VR por sindicato/estado diretamente do arquivo de configuração."""
        try:
            import pandas as pd
            
            # Caminho para o arquivo de configuração
            raiz_projeto = Path(__file__).resolve().parents[2]
            arquivo_valores = raiz_projeto / "input_data" / "configuracoes" / "Base sindicato x valor.xlsx"
            
            if not arquivo_valores.exists():
                logger.warning(f"Arquivo de valores de VR não encontrado: {arquivo_valores}. Não há valores padrão disponíveis!")
                # Ao invés de retornar valores hardcoded, retornamos um dicionário vazio
                # e forçamos o sistema a lidar com a ausência do arquivo
                return {}
                
            # Carregar o arquivo Excel
            df = pd.read_excel(arquivo_valores)
            
            # Criar dicionário com nome do estado como chave e valor de VR como valor
            valores_vr = {}
            for _, row in df.iterrows():
                estado = str(row["estado"]).strip() if "estado" in df.columns else ""
                valor = row["valor"] if "valor" in df.columns else None
                
                if estado and pd.notna(valor):
                    # Normalizar o nome do estado e garantir que é um Decimal
                    estado_normalizado = estado
                    valores_vr[estado_normalizado] = Decimal(str(valor))
            
            # Verificar se algum valor foi carregado
            if not valores_vr:
                logger.error(f"Nenhum valor de VR foi carregado do arquivo {arquivo_valores}. Verificar estrutura do arquivo!")
                # Retornar dicionário vazio e deixar o sistema lidar com a ausência de valores
                return {}
            
            logger.info(f"Valores de VR carregados: {len(valores_vr)} estados")
            
            # Log para depuração
            for estado, valor in valores_vr.items():
                logger.info(f"Valor VR carregado: {estado} = R$ {valor}")
                
            return valores_vr
            
        except Exception as e:
            logger.error(f"Erro ao carregar valores de VR: {e}. Não há valores padrão disponíveis!")
            # Ao invés de retornar valores hardcoded, lançamos uma exceção para forçar a correção do arquivo
            raise ValueError(f"Falha ao carregar valores de VR. Verifique o arquivo 'Base sindicato x valor.xlsx': {e}")
        
    def _carregar_tabela_dias_uteis(self) -> Dict[str, int]:
        """Carrega a tabela de dias úteis por sindicato do arquivo 'Base dias uteis.xlsx'."""
        try:
            import pandas as pd
            
            # Assume que este arquivo está em projeto_vr/passo_5_entrega_final/
            raiz_projeto = Path(__file__).resolve().parents[2]
            arquivo_dias_uteis = raiz_projeto / "input_data" / "configuracoes" / "Base dias uteis.xlsx"
            
            if not arquivo_dias_uteis.exists():
                logger.warning(f"Arquivo de dias úteis não encontrado: {arquivo_dias_uteis}. Usando valores padrão.")
                return {}
                
            # Carregar a tabela
            df = pd.read_excel(arquivo_dias_uteis)
            
            # Criar dicionário de sindicato -> dias úteis
            dias_por_sindicato = {}
            for _, row in df.iterrows():
                if pd.notna(row['sindicato']) and pd.notna(row['dias uteis']):
                    # Normalizar o nome do sindicato (remover espaços extras, converter para maiúsculas)
                    sindicato = str(row['sindicato']).strip().upper()
                    dias = int(row['dias uteis'])
                    dias_por_sindicato[sindicato] = dias
                    
            logger.info(f"Carregada tabela de dias úteis com {len(dias_por_sindicato)} sindicatos")
            return dias_por_sindicato
            
        except Exception as e:
            logger.error(f"Erro ao carregar tabela de dias úteis: {e}. Usando valores padrão.")
            return {}
        
    def _carregar_dias_uteis(self) -> Dict[str, int]:
        """Carrega os dias úteis por sindicato do arquivo de configuração."""
        try:
            import pandas as pd
            
            # Assume que o arquivo está em input_data/configuracoes/
            raiz_projeto = Path(__file__).resolve().parents[2]
            arquivo_dias_uteis = raiz_projeto / "input_data" / "configuracoes" / "Base dias uteis.xlsx"
            
            if not arquivo_dias_uteis.exists():
                logger.warning(f"Arquivo de dias úteis não encontrado: {arquivo_dias_uteis}")
                return {}
                
            # Carregar o arquivo Excel
            df = pd.read_excel(arquivo_dias_uteis)
            
            # Criar dicionário com nome do sindicato como chave e dias úteis como valor
            dias_uteis = {}
            for _, row in df.iterrows():
                # Extrair apenas o nome do sindicato sem o estado
                nome_completo = row['sindicato']
                dias = row['dias uteis']
                
                # Remover espaços extras e normalizar
                if isinstance(nome_completo, str):
                    nome_completo = nome_completo.strip()
                    dias_uteis[nome_completo] = int(dias)
                    
                    # Também adicionar versões simplificadas para facilitar a busca
                    if " - " in nome_completo:
                        sigla = nome_completo.split(" - ")[0].strip()
                        dias_uteis[sigla] = int(dias)
            
            logger.info(f"Dias úteis carregados para {len(dias_uteis)} sindicatos")
            return dias_uteis
            
        except Exception as e:
            logger.error(f"Erro ao carregar dias úteis: {e}")
            return {}

    def gerar_planilha_operadora(
        self, dados_validados: Dict[str, Any]
    ) -> Dict[str, str]:
        """Gera APENAS a planilha Excel final. Outros arquivos são temporários."""
        logger.info("Gerando planilha Excel final para operadora")

        arquivos_gerados = {}

        # ÚNICO ARQUIVO PERMANENTE: Planilha Excel
        try:
            arquivo_excel = self._gerar_excel_operadora(dados_validados)
            arquivos_gerados["planilha_excel"] = arquivo_excel
            logger.info("✅ Excel gerado com sucesso como ÚNICO output permanente")
        except ImportError as e:
            logger.error(f"❌ Erro ao gerar Excel: {e}")
            raise Exception("Excel é obrigatório. Instale pandas e openpyxl.")
        except Exception as e:
            logger.error(f"❌ Erro inesperado ao gerar Excel: {e}")
            raise

        # Arquivos temporários para controle (não ficam em output/)
        import os
        import tempfile

        # JSON temporário para debug
        temp_json = os.path.join(
            tempfile.gettempdir(),
            f"vr_dados_temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        )
        with open(temp_json, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "colaboradores": dados_validados.get("colaboradores", {}),
                    "metadata": dados_validados.get("metadata", {}),
                    "estatisticas": {
                        "total_colaboradores": len(
                            dados_validados.get("colaboradores", {})
                        ),
                        "data_geracao": datetime.now().isoformat(),
                    },
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
        logger.info(f"📄 Dados temporários salvos em: {temp_json}")

        # Relatório temporário
        temp_relatorio = os.path.join(
            tempfile.gettempdir(),
            f"vr_relatorio_temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
        )
        self._gerar_relatorio_temporario(dados_validados, temp_relatorio)
        logger.info(f"📋 Relatório temporário salvo em: {temp_relatorio}")

        logger.info(f"✅ ÚNICA saída permanente: {os.path.basename(arquivo_excel)}")
        return arquivos_gerados

    def _gerar_csv_operadora(self, dados: Dict[str, Any]) -> str:
        """Gera arquivo CSV no formato da operadora."""
        nome_arquivo = (
            f"VR_MENSAL_OPERADORA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        caminho_arquivo = self.diretorio_output / nome_arquivo

        logger.info(f"Gerando CSV: {caminho_arquivo}")

        # Cabeçalhos conforme modelo da operadora (exatamente como no modelo)
        cabecalhos = [
            "matricula",  # Matricula do colaborador
            "admissao",  # Data de admissão
            "sindicato",  # Sindicato do colaborador
            "competencia",  # Mês/ano de competência
            "dias",  # Dias úteis
            "valor diario",  # Valor diário do VR
            "TOTAL",  # Valor total VR
            "custo empresa",  # 80% do valor (empresa paga)
            "deconto funcionario",  # 20% do valor (desconto funcionário)
            "OBS GERAL",  # Observações gerais
        ]

        with open(caminho_arquivo, "w", newline="", encoding="utf-8") as arquivo:
            writer = csv.writer(
                arquivo, delimiter=";"
            )  # Usar ponto e vírgula como separador
            writer.writerow(cabecalhos)

            for matricula, colaborador in dados.get("colaboradores", {}).items():
                linha = self._preparar_linha_csv(matricula, colaborador)
                writer.writerow(linha)

        logger.info(f"CSV gerado com {len(dados.get('colaboradores', {}))} registros")
        return str(caminho_arquivo)

    def _preparar_linha_csv(self, matricula: str, colaborador: Dict) -> List[str]:
        """Prepara uma linha de dados para o CSV da operadora seguindo modelo exato."""
        # Obter valor VR de diferentes possíveis locais
        valor_vr = 0

        if "valor_vr_calculado" in colaborador and colaborador["valor_vr_calculado"]:
            valor_vr = Decimal(str(colaborador["valor_vr_calculado"]))
        elif "calculo_vr" in colaborador and isinstance(
            colaborador["calculo_vr"], dict
        ):
            calculo_vr = colaborador["calculo_vr"]
            if "valor_total" in calculo_vr and calculo_vr["valor_total"]:
                valor_vr = Decimal(str(calculo_vr["valor_total"]))

        # Obter dados do sindicato e dias úteis
        sindicato_info = colaborador.get("sindicato", {})
        sindicato_nome = ""
        dias_uteis = None  # Será definido após identificar corretamente o sindicato
        valor_diario = 0

        # Extrair nome do sindicato
        if isinstance(sindicato_info, dict):
            sindicato_nome = sindicato_info.get("nome", "")
        elif isinstance(sindicato_info, str):
            sindicato_nome = sindicato_info
            
        # Se o sindicato foi identificado, aplicar o número correto de dias úteis
        if sindicato_nome:
            # Normalizar o nome do sindicato para busca na tabela
            sindicato_normalizado = sindicato_nome.strip().upper()
            
            # Tentar obter os dias úteis da tabela de referência
            if sindicato_normalizado in self.tabela_dias_uteis:
                dias_uteis = self.tabela_dias_uteis[sindicato_normalizado]
                logger.info(f"Dias úteis para sindicato '{sindicato_nome}' encontrado na tabela: {dias_uteis}")
            else:
                # Se não encontrar pelo nome normalizado, tentar encontrar pelo estado
                estado_identificado = None
                for estado in ["São Paulo", "Rio Grande do Sul", "Rio de Janeiro", "Paraná"]:
                    if estado in sindicato_nome:
                        estado_identificado = estado
                        break
                        
                if estado_identificado:
                    # Mapear dias úteis por estado (conforme tabela de referência)
                    dias_por_estado = {
                        "São Paulo": 22,
                        "Paraná": 22,
                        "Rio Grande do Sul": 21,
                        "Rio de Janeiro": 21
                    }
                    dias_uteis = dias_por_estado.get(estado_identificado)
                    logger.info(f"Dias úteis para estado '{estado_identificado}' definido como: {dias_uteis}")
                else:
                    # Se ainda não encontrou, usar valor padrão
                    dias_uteis = 22
                    logger.warning(f"Estado não identificado para sindicato: {sindicato_nome}. Usando dias úteis padrão: {dias_uteis}")
        else:
            # Sindicato não identificado, usar valor padrão
            dias_uteis = 22
            logger.warning(f"Sindicato não identificado para matrícula {matricula}. Usando dias úteis padrão: {dias_uteis}")
                
        # Registrar informação sobre dias úteis para debugging
        logger.info(f"Sindicato: {sindicato_nome}, Dias úteis: {dias_uteis}")
        
        # AJUSTE PARA RESPEITAR OS VALORES ESPECÍFICOS DE CADA SINDICATO
        # E AINDA ATINGIR O VALOR TOTAL DE 1.380.178,00 (USANDO FATOR DE CORREÇÃO)
        # 
        # Removido o fator de correção para garantir que os valores exatos das CCTs sejam respeitados.
        # Os valores diários a seguir são os valores exatos especificados nas Convenções Coletivas
        # de Trabalho de cada estado, sem nenhum tipo de ajuste ou correção.
        # 
        # Valores oficiais das CCTs:
        # - São Paulo: R$ 37,50 (CCT-2024_2025- São Paulo)
        # - Rio Grande do Sul: R$ 35,00 (CCT-2024_2025- Rio Grande do Sul)
        # - Rio de Janeiro: R$ 35,00 (CCT2023-2025-1 Rio de Janeiro)
        # - Paraná: R$ 38,00 conforme CCT 2025-2027 atualizada
        # 
        # Esses valores são aplicados diretamente, sem ajustes ou fatores de correção.

        # Ajuste proporcional para qualquer situação (férias, afastamento, admissão/desligamento, etc)
        dias_ferias = colaborador.get("dias_ferias", 0)
        dias_afastados = colaborador.get("dias_afastados", 0)
        dias_admissao = colaborador.get("dias_admissao", 0)
        dias_desligamento = colaborador.get("dias_desligamento", 0)
        situacao = colaborador.get("situacao", "").lower()
        
        # Forçar cálculo proporcional para férias, mesmo que não tenha dias_ferias informados
        if "féria" in situacao and dias_ferias == 0:
            dias_ferias = 15  # Valor padrão se não informado
            logger.info(f"Férias padrão (15 dias) aplicado para matrícula {matricula} com situação: {situacao}")
            
        # Calcular proporcional para admissões recentes
        # Verificar se é um colaborador admitido em abril
        admissao = colaborador.get("admissao", "")
        if admissao and isinstance(admissao, str) and admissao.startswith("2025-04"):
            # Calcular dias não trabalhados em abril com base na data de admissão
            try:
                from datetime import datetime
                data_admissao = datetime.strptime(admissao, "%Y-%m-%d")
                # Período de referência: 15/04 a 15/05
                inicio_mes = datetime(2025, 4, 1)
                # Usar os dias úteis específicos do sindicato do colaborador
                dias_uteis_abril = dias_uteis  # Usar o valor já calculado para este sindicato
                dia_admissao = data_admissao.day
                proporcao_nao_trabalhada = (dia_admissao - 1) / 30  # -1 porque no dia da admissão já trabalha
                dias_admissao = int(proporcao_nao_trabalhada * dias_uteis_abril)
                logger.info(f"Admissão proporcional para matrícula {matricula}: data={admissao}, dias não trabalhados={dias_admissao}, dias úteis total={dias_uteis_abril}")
            except Exception as e:
                logger.error(f"Erro ao calcular proporcional de admissão para {matricula}: {e}")
                dias_admissao = 0
        
        # Calcular dias trabalhados considerando todos os tipos de afastamento/admissão/desligamento
        dias_trabalhados = dias_uteis
        if dias_ferias > 0:
            dias_trabalhados -= dias_ferias
        if dias_afastados > 0:
            dias_trabalhados -= dias_afastados
        if dias_admissao > 0:
            dias_trabalhados -= dias_admissao
        if dias_desligamento > 0:
            dias_trabalhados -= dias_desligamento
        if dias_trabalhados < 0:
            dias_trabalhados = 0
            
        # Verificação explícita para situações onde não deve haver pagamento de VR
        # conforme regras da empresa e legislação
        situacao_lower = situacao.lower() if situacao else ""
        if (situacao_lower == "licença maternidade" or 
            situacao_lower == "auxílio doença" or
            situacao_lower == "exterior" or 
            "aprendiz" in situacao_lower or 
            "estagiário" in situacao_lower):
            # Registrar explicitamente o motivo da exclusão
            observacoes = f"VR não aplicável: {situacao}"
            dias_trabalhados = 0
        
        # Usar os valores EXATOS específicos por sindicato conforme as CCTs atualizadas
        # Estes valores são carregados dinamicamente do arquivo de configuração "Base sindicato x valor.xlsx"
        # Não aplicamos fator de correção para manter os valores exatos das CCTs
        # O cálculo é feito com base em Decimal para maior precisão
        
        # Verificar se temos valores de VR carregados
        if not self.valores_vr_sindicato:
            raise ValueError("Nenhum valor de VR encontrado! Verifique o arquivo 'Base sindicato x valor.xlsx'")
        
        # Identificar o estado correto com base no nome do sindicato
        estado_identificado = None
        
        # Mapeamento detalhado dos sindicatos para os estados
        mapeamento_sindicatos = {
            "SINDPD SP": "São Paulo",
            "SP": "São Paulo",
            "SIND.TRAB.EM PROC DADOS E EMPR.EMPRESAS PROC DADOS ESTADO DE SP": "São Paulo",
            
            "SINDPPD RS": "Rio Grande do Sul",
            "RS": "Rio Grande do Sul",
            "SINDICATO DOS TRAB. EM PROC. DE DADOS RIO GRANDE DO SUL": "Rio Grande do Sul",
            
            "SINDPD RJ": "Rio de Janeiro",
            "RJ": "Rio de Janeiro",
            "SINDICATO PROFISSIONAIS DE PROC DADOS DO RIO DE JANEIRO": "Rio de Janeiro",
            
            "SITEPD PR": "Paraná",
            "PR": "Paraná",
            "SIND DOS TRAB EM EMPR PRIVADAS DE PROC DE DADOS DE CURITIBA": "Paraná"
        }
        
        # Primeiro tentamos uma correspondência exata com o nome do sindicato normalizado
        sindicato_upper = sindicato_nome.upper()
        
        for key, estado in mapeamento_sindicatos.items():
            if key.upper() in sindicato_upper:
                estado_identificado = estado
                logger.info(f"Estado identificado por correspondência de sindicato: {estado} para {sindicato_nome}")
                break
        
        # Se não encontrou por nome do sindicato, tenta por código de estado ou nome do estado
        if estado_identificado is None:
            # Verificar pelo nome completo do estado
            for estado in self.valores_vr_sindicato.keys():
                if estado in sindicato_nome:
                    estado_identificado = estado
                    logger.info(f"Estado identificado pelo nome: {estado}")
                    break
            
            # Se ainda não encontrou, tenta identificar por códigos de estado
            if estado_identificado is None:
                # Mapeamento de palavras-chave para estados
                keywords_estados = {
                    "SP": "São Paulo",
                    "RS": "Rio Grande do Sul",
                    "RJ": "Rio de Janeiro",
                    "PR": "Paraná"
                }
                
                # Verificar se alguma das palavras-chave está presente no nome do sindicato
                for keyword, estado in keywords_estados.items():
                    if keyword in sindicato_nome and estado in self.valores_vr_sindicato:
                        estado_identificado = estado
                        logger.info(f"Estado identificado pelo código {keyword}: {estado}")
                        break
        
        # Usar o valor correspondente do estado identificado ou buscar um valor padrão
        if estado_identificado and estado_identificado in self.valores_vr_sindicato:
            valor_base = self.valores_vr_sindicato[estado_identificado]
            logger.info(f"Usando valor de {estado_identificado}: R$ {valor_base} para sindicato: {sindicato_nome}")
        else:
            # Se o sindicato contém "PR" mas o estado não foi identificado, forçar valor do Paraná
            if "PR" in sindicato_nome and "Paraná" in self.valores_vr_sindicato:
                estado_identificado = "Paraná"
                valor_base = self.valores_vr_sindicato["Paraná"]
                logger.info(f"Forçando uso do valor do Paraná: R$ {valor_base} para sindicato contendo 'PR': {sindicato_nome}")
            # Se o sindicato contém "SP" mas o estado não foi identificado, forçar valor de São Paulo
            elif "SP" in sindicato_nome and "São Paulo" in self.valores_vr_sindicato:
                estado_identificado = "São Paulo"
                valor_base = self.valores_vr_sindicato["São Paulo"]
                logger.info(f"Forçando uso do valor de São Paulo: R$ {valor_base} para sindicato contendo 'SP': {sindicato_nome}")
            # Se o sindicato contém "RS" mas o estado não foi identificado, forçar valor do Rio Grande do Sul
            elif "RS" in sindicato_nome and "Rio Grande do Sul" in self.valores_vr_sindicato:
                estado_identificado = "Rio Grande do Sul"
                valor_base = self.valores_vr_sindicato["Rio Grande do Sul"]
                logger.info(f"Forçando uso do valor do Rio Grande do Sul: R$ {valor_base} para sindicato contendo 'RS': {sindicato_nome}")
            # Se o sindicato contém "RJ" mas o estado não foi identificado, forçar valor do Rio de Janeiro
            elif "RJ" in sindicato_nome and "Rio de Janeiro" in self.valores_vr_sindicato:
                estado_identificado = "Rio de Janeiro"
                valor_base = self.valores_vr_sindicato["Rio de Janeiro"]
                logger.info(f"Forçando uso do valor do Rio de Janeiro: R$ {valor_base} para sindicato contendo 'RJ': {sindicato_nome}")
            else:
                # Para colaboradores sem sindicato ou estado identificado, usar o valor do Paraná como padrão
                # já que este é o valor correto da CCT atualizada (2025-2027)
                if "Paraná" in self.valores_vr_sindicato:
                    valor_base = self.valores_vr_sindicato["Paraná"]
                    logger.warning(f"Estado não identificado para sindicato: {sindicato_nome}. Usando valor do Paraná como padrão: R$ {valor_base}")
                else:
                    # Encontrar um valor padrão nos dados carregados, sem hardcoding
                    estados_disponiveis = list(self.valores_vr_sindicato.keys())
                    if estados_disponiveis:
                        # Usar o primeiro estado disponível como fallback
                        estado_padrao = estados_disponiveis[0]
                        valor_base = self.valores_vr_sindicato[estado_padrao]
                        logger.warning(f"Estado não identificado para sindicato: {sindicato_nome}. Usando valor de {estado_padrao}: R$ {valor_base}")
                    else:
                        # Se não houver nenhum estado disponível, lançar erro
                        raise ValueError("Nenhum valor de VR disponível para cálculo!")
        
        # Usamos o valor base exato da CCT, sem fator de correção
        valor_diario = valor_base
            
        # SEMPRE calcular valor proporcional baseado nos dias trabalhados
        # Não usar valor cheio em nenhum caso
        total_vr = valor_diario * Decimal(str(dias_trabalhados))
        valor_empresa = total_vr * Decimal(str(self.custo_empresa_percentual))
        valor_funcionario = total_vr * Decimal(str(self.custo_funcionario_percentual))

        # Formatar datas conforme modelo (YYYY-MM-DD para Excel processar corretamente)
        def formatar_data_excel(data_valor):
            if not data_valor:
                return ""
            if isinstance(data_valor, (date, datetime)):
                return data_valor.strftime("%Y-%m-%d")
            if isinstance(data_valor, str):
                # Tentar converter diferentes formatos
                for formato in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
                    try:
                        data_obj = datetime.strptime(data_valor, formato)
                        return data_obj.strftime("%Y-%m-%d")
                    except ValueError:
                        continue
            return str(data_valor)

        # Competência (primeiro dia do mês de referência)
        competencia = "2025-05-01"  # Maio 2025 conforme modelo

        # Observações baseadas na situação
        observacoes = ""
        situacao = colaborador.get("situacao", "")
        
        # Documentação detalhada para valores atípicos e situações especiais
        if situacao and situacao.lower() != "trabalhando":
            observacoes = f"Situação: {situacao}"
            
        # Adicionar justificativa para dias parciais
        if dias_trabalhados > 0 and dias_trabalhados < dias_uteis:
            motivo = []
            if dias_ferias > 0:
                motivo.append(f"{dias_ferias} dias em férias")
            if dias_afastados > 0:
                motivo.append(f"{dias_afastados} dias afastado")
            if dias_admissao > 0:
                motivo.append(f"Admitido há {dias_admissao} dias")
            if dias_desligamento > 0:
                motivo.append(f"Desligado há {dias_desligamento} dias")
                
            if motivo:
                observacoes = f"VR proporcional: {', '.join(motivo)}"

        # Corrigir campo 'dias' para mostrar dias trabalhados sempre que houver cálculo proporcional
        # Para situação de férias ou outro afastamento, sempre usar dias_trabalhados
        situacao = colaborador.get("situacao", "").lower()
        dias_para_planilha = dias_trabalhados  # Sempre usar dias trabalhados, não o total
        
        # Log para depuração dos dias úteis e trabalhados
        logger.debug(f"Colaborador {matricula} - Sindicato: {sindicato_nome} - Dias úteis: {dias_uteis} - Dias trabalhados: {dias_trabalhados}")
        return [
            matricula,  # matricula
            formatar_data_excel(colaborador.get("admissao", "")),  # admissao
            sindicato_nome,  # sindicato
            competencia,  # competencia
            dias_para_planilha if dias_para_planilha is not None else None,  # dias
            float(valor_diario) if valor_diario else None,  # valor diario
            float(total_vr) if total_vr else None,  # TOTAL
            float(valor_empresa) if valor_empresa else None,  # custo empresa
            float(valor_funcionario) if valor_funcionario else None,  # deconto funcionario
            observacoes,  # OBS GERAL
        ]

    def _formatar_cpf(self, cpf: str) -> str:
        """Formata CPF no padrão XXX.XXX.XXX-XX."""
        if not cpf:
            return ""

        # Remover formatação existente
        cpf_numeros = "".join(filter(str.isdigit, str(cpf)))

        if len(cpf_numeros) == 11:
            return f"{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}"

        return str(cpf)

    def _gerar_json_operadora(self, dados: Dict[str, Any]) -> str:
        """Gera arquivo JSON estruturado para a operadora."""
        nome_arquivo = (
            f"VR_MENSAL_DADOS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        caminho_arquivo = self.diretorio_output / nome_arquivo

        logger.info(f"Gerando JSON: {caminho_arquivo}")

        # Preparar dados estruturados
        dados_operadora = {
            "cabecalho": {
                "empresa": "I2A2 TECNOLOGIA",
                "periodo_referencia": "15/04/2025 a 15/05/2025",
                "data_geracao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "total_colaboradores": len(dados.get("colaboradores", {})),
                "valor_total_vr": 0,
                "valor_total_empresa": 0,
                "valor_total_funcionario": 0,
            },
            "colaboradores": [],
            "resumo": {
                "distribuicao_por_estado": {},
                "distribuicao_por_situacao": {},
                "estatisticas_valores": {},
            },
        }

        # Processar colaboradores
        total_vr = Decimal("0")
        total_empresa = Decimal("0")
        total_funcionario = Decimal("0")

        distribuicao_estado = {}
        distribuicao_situacao = {}

        for matricula, colaborador in dados.get("colaboradores", {}).items():
            # Obter valor VR de diferentes possíveis locais
            valor_vr = 0

            if (
                "valor_vr_calculado" in colaborador
                and colaborador["valor_vr_calculado"]
            ):
                valor_vr = Decimal(str(colaborador["valor_vr_calculado"]))
            elif "calculo_vr" in colaborador and isinstance(
                colaborador["calculo_vr"], dict
            ):
                calculo_vr = colaborador["calculo_vr"]
                if "valor_total" in calculo_vr and calculo_vr["valor_total"]:
                    valor_vr = Decimal(str(calculo_vr["valor_total"]))

            # Calcular valores empresa e funcionário
            valor_empresa = valor_vr * Decimal(str(self.custo_empresa_percentual))
            valor_funcionario = valor_vr * Decimal(
                str(self.custo_funcionario_percentual)
            )

            # Somar totais
            total_vr += valor_vr
            total_empresa += valor_empresa
            total_funcionario += valor_funcionario

            # Contabilizar distribuições
            estado = colaborador.get("endereco", {}).get("estado", "Não informado")
            situacao = colaborador.get("situacao", "Trabalhando")

            distribuicao_estado[estado] = distribuicao_estado.get(estado, 0) + 1
            distribuicao_situacao[situacao] = distribuicao_situacao.get(situacao, 0) + 1

            # Preparar dados do colaborador
            dados_colaborador = {
                "matricula": matricula,
                "nome": colaborador.get("nome", ""),
                "cpf": colaborador.get("cpf", ""),
                "valores": {
                    "vr_total": float(valor_vr),
                    "empresa_80pct": float(valor_empresa),
                    "funcionario_20pct": float(valor_funcionario),
                },
                "vigencia": {
                    "inicio": colaborador.get("data_inicio_vigencia", "15/04/2025"),
                    "fim": colaborador.get("data_fim_vigencia", "15/05/2025"),
                },
                "dados_funcionais": {
                    "empresa": colaborador.get("empresa", ""),
                    "cargo": colaborador.get("cargo", ""),
                    "situacao": situacao,
                    "admissao": colaborador.get("admissao", ""),
                    "demissao": colaborador.get("demissao", ""),
                },
                "endereco": colaborador.get("endereco", {}),
            }

            dados_operadora["colaboradores"].append(dados_colaborador)

        # Atualizar totais no cabeçalho
        dados_operadora["cabecalho"]["valor_total_vr"] = float(total_vr)
        dados_operadora["cabecalho"]["valor_total_empresa"] = float(total_empresa)
        dados_operadora["cabecalho"]["valor_total_funcionario"] = float(
            total_funcionario
        )

        # Atualizar resumo
        dados_operadora["resumo"]["distribuicao_por_estado"] = distribuicao_estado
        dados_operadora["resumo"]["distribuicao_por_situacao"] = distribuicao_situacao
        dados_operadora["resumo"]["estatisticas_valores"] = {
            "valor_medio_vr": (
                float(total_vr / len(dados.get("colaboradores", {})))
                if dados.get("colaboradores")
                else 0
            ),
            "maior_valor": max(
                [
                    float(c.get("valor_vr_calculado", 0))
                    for c in dados.get("colaboradores", {}).values()
                ],
                default=0,
            ),
            "menor_valor": min(
                [
                    float(c.get("valor_vr_calculado", 0))
                    for c in dados.get("colaboradores", {}).values()
                    if float(c.get("valor_vr_calculado", 0)) > 0
                ],
                default=0,
            ),
        }

        # Salvar JSON
        with open(caminho_arquivo, "w", encoding="utf-8") as arquivo:
            json.dump(
                dados_operadora, arquivo, indent=2, ensure_ascii=False, default=str
            )

        logger.info(
            f"JSON gerado com dados de {len(dados_operadora['colaboradores'])} colaboradores"
        )
        return str(caminho_arquivo)

    def _gerar_relatorio_controle(self, dados: Dict[str, Any]) -> str:
        """Gera relatório de controle para conferência."""
        nome_arquivo = f"RELATORIO_CONTROLE_OPERADORA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        caminho_arquivo = self.diretorio_output / nome_arquivo

        logger.info(f"Gerando relatório de controle: {caminho_arquivo}")

        relatorio = []
        relatorio.append("=" * 80)
        relatorio.append("RELATÓRIO DE CONTROLE - ENTREGA OPERADORA VR")
        relatorio.append("=" * 80)
        relatorio.append(
            f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        )
        relatorio.append(f"Período de referência: 15/04/2025 a 15/05/2025")
        relatorio.append("")

        # Estatísticas gerais
        colaboradores = dados.get("colaboradores", {})
        total_colaboradores = len(colaboradores)

        relatorio.append("ESTATÍSTICAS GERAIS:")
        relatorio.append("-" * 40)
        relatorio.append(f"Total de colaboradores: {total_colaboradores}")

        # Calcular totais
        total_vr = Decimal("0")
        total_empresa = Decimal("0")
        total_funcionario = Decimal("0")

        for colaborador in colaboradores.values():
            # Obter valor VR de diferentes possíveis locais
            valor_vr = 0

            if (
                "valor_vr_calculado" in colaborador
                and colaborador["valor_vr_calculado"]
            ):
                valor_vr = Decimal(str(colaborador["valor_vr_calculado"]))
            elif "calculo_vr" in colaborador and isinstance(
                colaborador["calculo_vr"], dict
            ):
                calculo_vr = colaborador["calculo_vr"]
                if "valor_total" in calculo_vr and calculo_vr["valor_total"]:
                    valor_vr = Decimal(str(calculo_vr["valor_total"]))

            total_vr += valor_vr
            total_empresa += valor_vr * Decimal(str(self.custo_empresa_percentual))
            total_funcionario += valor_vr * Decimal(
                str(self.custo_funcionario_percentual)
            )

        relatorio.append(f"Valor total VR: R$ {total_vr:,.2f}")
        relatorio.append(f"Valor total empresa (80%): R$ {total_empresa:,.2f}")
        relatorio.append(f"Valor total funcionário (20%): R$ {total_funcionario:,.2f}")
        relatorio.append(
            f"Valor médio por colaborador: R$ {total_vr/total_colaboradores if total_colaboradores > 0 else 0:,.2f}"
        )
        relatorio.append("")

        # Distribuição por estado
        distribuicao_estado = {}
        distribuicao_situacao = {}

        for colaborador in colaboradores.values():
            estado = colaborador.get("endereco", {}).get("estado", "Não informado")
            situacao = colaborador.get("situacao", "Trabalhando")

            if estado not in distribuicao_estado:
                distribuicao_estado[estado] = {"count": 0, "valor": Decimal("0")}
            if situacao not in distribuicao_situacao:
                distribuicao_situacao[situacao] = {"count": 0, "valor": Decimal("0")}

            valor_vr = Decimal(str(colaborador.get("valor_vr_calculado", 0)))
            distribuicao_estado[estado]["count"] += 1
            distribuicao_estado[estado]["valor"] += valor_vr
            distribuicao_situacao[situacao]["count"] += 1
            distribuicao_situacao[situacao]["valor"] += valor_vr

        relatorio.append("DISTRIBUIÇÃO POR ESTADO:")
        relatorio.append("-" * 40)
        for estado, dados in sorted(distribuicao_estado.items()):
            relatorio.append(
                f"{estado}: {dados['count']} colaboradores - R$ {dados['valor']:,.2f}"
            )
        relatorio.append("")

        relatorio.append("DISTRIBUIÇÃO POR SITUAÇÃO:")
        relatorio.append("-" * 40)
        for situacao, dados in sorted(distribuicao_situacao.items()):
            relatorio.append(
                f"{situacao}: {dados['count']} colaboradores - R$ {dados['valor']:,.2f}"
            )
        relatorio.append("")

        # Validações aplicadas
        validacao = dados.get("validacao", {})
        if validacao:
            estatisticas = validacao.get("estatisticas", {})
            relatorio.append("VALIDAÇÕES APLICADAS:")
            relatorio.append("-" * 40)
            relatorio.append(
                f"Registros processados: {estatisticas.get('total_registros', 0)}"
            )
            relatorio.append(
                f"Registros válidos: {estatisticas.get('registros_validos', 0)}"
            )
            relatorio.append(
                f"Registros com erro: {estatisticas.get('registros_com_erro', 0)}"
            )
            relatorio.append(
                f"Registros com warning: {estatisticas.get('registros_com_warning', 0)}"
            )

            if estatisticas.get("total_registros", 0) > 0:
                taxa_aprovacao = (
                    estatisticas.get("registros_validos", 0)
                    / estatisticas["total_registros"]
                ) * 100
                relatorio.append(f"Taxa de aprovação: {taxa_aprovacao:.2f}%")
            relatorio.append("")

        # Instruções para a operadora
        relatorio.append("INSTRUÇÕES PARA OPERADORA:")
        relatorio.append("-" * 40)
        relatorio.append("1. Conferir total de colaboradores")
        relatorio.append("2. Validar soma dos valores (VR = Empresa + Funcionário)")
        relatorio.append("3. Verificar formatação de datas (DD/MM/YYYY)")
        relatorio.append("4. Confirmar CPFs formatados (XXX.XXX.XXX-XX)")
        relatorio.append("5. Validar período de vigência (15/04/2025 a 15/05/2025)")
        relatorio.append("")

        relatorio.append("OBSERVAÇÕES:")
        relatorio.append("-" * 40)
        relatorio.append("• Valor empresa: 80% do VR total")
        relatorio.append("• Valor funcionário: 20% do VR total (desconto em folha)")
        relatorio.append("• Apenas colaboradores ativos estão inclusos")
        relatorio.append("• Cálculo baseado em dias úteis do período")
        relatorio.append("")

        relatorio.append("=" * 80)

        # Salvar relatório
        with open(caminho_arquivo, "w", encoding="utf-8") as arquivo:
            arquivo.write("\n".join(relatorio))

        logger.info("Relatório de controle gerado")
        return str(caminho_arquivo)

    def _gerar_excel_operadora(self, dados: Dict[str, Any]) -> str:
        """Gera arquivo Excel no formato da operadora (usando openpyxl)."""
        try:
            import openpyxl
            import pandas as pd
        except ImportError:
            raise ImportError("Pandas e openpyxl necessários para geração de Excel")

        nome_arquivo = "VR MENSAL 05.2025.xlsx"
        caminho_arquivo = self.diretorio_output / nome_arquivo
        
        # Valor esperado conforme especificação
        valor_esperado = Decimal('1380178.00')
        
        logger.info(f"Gerando Excel: {caminho_arquivo}")

        # Preparar dados para DataFrame
        dados_planilha = []

        for matricula, colaborador in dados.get("colaboradores", {}).items():
            linha = self._preparar_linha_csv(matricula, colaborador)

            # Converter valores monetários para float (exceto OBS GERAL)
            for i in [5, 6, 7, 8]:  # Índices dos valores monetários
                if i < len(linha) and linha[i]:
                    valor_str = str(linha[i]).replace(",", ".")
                    try:
                        linha[i] = float(valor_str)
                    except ValueError:
                        linha[i] = None
            # OBS GERAL sempre texto ou vazio
            if len(linha) > 9 and (linha[9] == 0 or linha[9] is None):
                linha[9] = ""
                
            # Verificar se os dias estão corretos e valor diário correto (debugging)
            logger.debug(f"Colaborador {linha[0]} - Situação: {colaborador.get('situacao', '')} - Dias: {linha[4]} - Valor diário: {linha[5]}")

            dados_planilha.append(linha)

        # Criar DataFrame
        colunas = [
            "matricula",
            "admissao",
            "sindicato",
            "competencia",
            "dias",
            "valor diario",
            "TOTAL",
            "custo empresa",
            "deconto funcionario",
            "OBS GERAL"
        ]

        df = pd.DataFrame(dados_planilha, columns=colunas)
        
        # Verificar o valor total e validar se está dentro do esperado
        valor_total_calculado = df['TOTAL'].sum()
        diferenca = abs(valor_total_calculado - float(valor_esperado))
        percentual_diferenca = (diferenca / float(valor_esperado)) * 100
        
        logger.info(f"Valor total calculado: R$ {valor_total_calculado:,.2f}")
        logger.info(f"Valor esperado: R$ {float(valor_esperado):,.2f}")
        logger.info(f"Diferença: R$ {diferenca:,.2f} ({percentual_diferenca:.4f}%)")
        
        # Validar se a diferença é aceitável (menos de 0.1%)
        if percentual_diferenca < 0.05:
            logger.info("✅ Valor total EXCELENTE - dentro da tolerância (<0.05%)")
        elif percentual_diferenca < 0.1:
            logger.info("✅ Valor total dentro da tolerância aceitável (< 0.1%)")
        else:
            logger.warning(f"⚠️ Diferença entre valor calculado e esperado: {percentual_diferenca:.4f}%")

        # Salvar Excel usando openpyxl
        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl") as writer:
            # Salvar somente a planilha principal, sem guias de documentação e validação
            df.to_excel(writer, sheet_name="VR MENSAL 05.2025", index=False)
            
            # Obter worksheet para formatação
            worksheet = writer.sheets["VR MENSAL 05.2025"]

            # Formatar cabeçalhos
            from openpyxl.styles import Alignment, Font, PatternFill

            # Estilo do cabeçalho
            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill(
                start_color="D7E4BC", end_color="D7E4BC", fill_type="solid"
            )
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Aplicar estilo no cabeçalho da planilha principal
            for col in range(1, len(colunas) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Ajustar largura das colunas conforme modelo
            column_widths = {
                "A": 12,  # matricula
                "B": 15,  # admissao
                "C": 35,  # sindicato
                "D": 15,  # competencia
                "E": 8,   # dias
                "F": 12,  # valor diario
                "G": 12,  # TOTAL
                "H": 15,  # custo empresa
                "I": 18,  # deconto funcionario
                "J": 25,  # OBS GERAL
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

        logger.info(f"Excel gerado com {len(dados_planilha)} registros")
        return str(caminho_arquivo)

    def _gerar_relatorio_temporario(self, dados: Dict[str, Any], caminho_arquivo: str):
        """Gera relatório temporário de controle."""
        colaboradores = dados.get("colaboradores", {})
        total_colaboradores = len(colaboradores)

        # Calcular estatísticas
        valor_total = 0
        valor_empresa_total = 0
        valor_funcionario_total = 0

        for colaborador in colaboradores.values():
            if (
                "valor_vr_calculado" in colaborador
                and colaborador["valor_vr_calculado"]
            ):
                vr = float(colaborador["valor_vr_calculado"])
            elif "calculo_vr" in colaborador and isinstance(
                colaborador["calculo_vr"], dict
            ):
                vr = float(colaborador["calculo_vr"].get("valor_total", 0))
            else:
                vr = 0

            valor_total += vr
            valor_empresa_total += vr * self.custo_empresa_percentual
            valor_funcionario_total += vr * self.custo_funcionario_percentual

        # Gerar relatório
        relatorio = []
        relatorio.append("=" * 60)
        relatorio.append("RELATÓRIO TEMPORÁRIO - VR OPERADORA")
        relatorio.append("=" * 60)
        relatorio.append(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        relatorio.append("")
        relatorio.append("RESUMO EXECUTIVO:")
        relatorio.append(f"• Total colaboradores: {total_colaboradores:,}")
        relatorio.append(f"• Valor total VR: R$ {valor_total:,.2f}")
        relatorio.append(f"• Custo empresa (80%): R$ {valor_empresa_total:,.2f}")
        relatorio.append(
            f"• Desconto funcionário (20%): R$ {valor_funcionario_total:,.2f}"
        )
        relatorio.append("")
        relatorio.append("STATUS: ✅ Excel gerado com sucesso")
        relatorio.append("=" * 60)

        with open(caminho_arquivo, "w", encoding="utf-8") as arquivo:
            arquivo.write("\n".join(relatorio))


def main():
    """Função principal para testar o gerador."""
    logger.info("📊 TESTANDO GERADOR DE PLANILHA FINAL")
    logger.info("=" * 50)

    # Dados de teste
    dados_teste = {
        "metadata": {"data_processamento": datetime.now().isoformat()},
        "colaboradores": {
            "12345": {
                "nome": "João Silva",
                "cpf": "12345678901",
                "valor_vr_calculado": 750.00,
                "empresa": "1410",
                "cargo": "ANALISTA",
                "situacao": "Trabalhando",
                "admissao": "2024-01-15",
                "endereco": {"estado": "São Paulo", "municipio": "São Paulo"},
            },
            "67890": {
                "nome": "Maria Santos",
                "cpf": "98765432100",
                "valor_vr_calculado": 825.00,
                "empresa": "1410",
                "cargo": "COORDENADOR",
                "situacao": "Trabalhando",
                "admissao": "2023-05-20",
                "endereco": {"estado": "Rio de Janeiro", "municipio": "Rio de Janeiro"},
            },
        },
    }

    gerador = GeradorPlanilhaFinal()
    arquivos = gerador.gerar_planilha_operadora(dados_teste)

    logger.info("📁 ARQUIVOS GERADOS:")
    for tipo, arquivo in arquivos.items():
        logger.info(f"  {tipo}: {arquivo}")

    logger.info("✅ Geração concluída!")


if __name__ == "__main__":
    main()
